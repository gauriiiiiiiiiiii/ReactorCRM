"""Export leads to CSV or Excel."""
from __future__ import annotations

import os
import json
from datetime import datetime

import pandas as pd
from database.models import Lead, TrackedPost


def export_leads(post_id: int | None, fmt: str, exports_dir: str) -> str:
    """
    Export leads for a post (or all posts if post_id is None).
    Returns the file path of the exported file.
    """
    query = Lead.query
    if post_id:
        query = query.filter_by(post_id=post_id)

    leads = query.order_by(Lead.score.desc()).all()
    rows = []
    for lead in leads:
        breakdown = {}
        try:
            breakdown = json.loads(lead.score_breakdown or "{}")
        except Exception:
            pass

        post_url = ""
        if lead.post:
            post_url = lead.post.url

        rows.append({
            "Name": lead.name,
            "Headline": lead.headline,
            "Location": lead.location,
            "Profile URL": lead.profile_url,
            "Company": lead.company,
            "Industry": lead.industry,
            "Email": lead.email,
            "Engagement Type": lead.engagement_type,
            "Reaction Type": lead.reaction_type,
            "Comment": lead.comment_text,
            "Lead Score": lead.score,
            "Score Tier": _tier(lead.score),
            "Status": lead.status,
            "Post URL": post_url,
            "Extracted At": lead.created_at.strftime("%Y-%m-%d %H:%M") if lead.created_at else "",
        })

    df = pd.DataFrame(rows)
    os.makedirs(exports_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = f"post{post_id}" if post_id else "all"

    if fmt == "excel":
        path = os.path.join(exports_dir, f"leads_{suffix}_{timestamp}.xlsx")
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Leads")
            _style_excel(writer, df)
    else:
        path = os.path.join(exports_dir, f"leads_{suffix}_{timestamp}.csv")
        df.to_csv(path, index=False)

    return path


def _tier(score: float) -> str:
    if score >= 70:
        return "Hot"
    if score >= 45:
        return "Warm"
    return "Cold"


def _style_excel(writer, df):
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        ws = writer.sheets["Leads"]

        # Header row
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0A66C2")
            cell.alignment = Alignment(horizontal="center")

        # Color-code score tiers
        score_col = None
        for idx, col in enumerate(df.columns, 1):
            if col == "Score Tier":
                score_col = idx
                break

        if score_col:
            tier_colors = {"Hot": "FF4C4C", "Warm": "FFA500", "Cold": "4CAF50"}
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                tier_cell = row[score_col - 1]
                color = tier_colors.get(tier_cell.value, "FFFFFF")
                tier_cell.fill = PatternFill("solid", fgColor=color)
                tier_cell.font = Font(bold=True, color="FFFFFF")

        # Auto-width columns
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    except Exception:
        pass
